from openpyxl import load_workbook


def path_planilha(nome_planilha):
    """
        Retornar o caminho completo da planilha

    :param nome_planilha: str com o nome da planilha
    :return: o caminho completo da planilha
    """
    return f'G:\\AmbientePython\\PycharmProjects\\gppinovacao\\{nome_planilha}'


def abrir_planilha(nome_planilha):
    """
       Abrir uma planilha excel existente

       :param nome_planilha: str com o nome da planilha
       :return: um objeto workbook ativado
       """
    # nome_planilha = f'G:\\AmbientePython\\PycharmProjects\\gppinovacao\\{nome_planilha}'
    return load_workbook(path_planilha(nome_planilha)).active


def mostrar_conteudo_planilha(planilha):
    qtdeMaximaLinha = planilha.max_row
    qtdeMaximaColuna = planilha.max_column
    for i in range(1, qtdeMaximaLinha + 1):
        lista = []
        for j in range(1, qtdeMaximaColuna + 1):
            lista.append(planilha.cell(row=i, column=j).value)
        print(lista)


def salvar_planilha(wb, path):
    wb.save(path)


if __name__ == '__main__':

    planilha = abrir_planilha('Contato.xlsx')

    # Salvar conteudo da planilha
    print('\nSalvar conteudo da planilha')
    path = path_planilha('Contato.xlsx')
    wb = load_workbook(path, False)
    folha = wb['Plan1']
    print(folha['A2'].value)
    folha['A2'] = 'Osvaldo Gonzaga Pires'
    salvar_planilha(wb, path)
    wb.close()

    # mostrar o título da planilha
    # planilha = abrir_planilha('gppInovacao.xlsx')
    # print(type(planilha))
    print('\nMostrar titulo da planilha')
    print(planilha.title)
    print(planilha.dimensions)
    print(type(planilha))

    # percorrer as folhas da planilha
    path = path_planilha('Contato.xlsx')
    folhas = load_workbook(path).sheetnames
    for folha in folhas:
        print(f'\nMostrar conteudos da folha {folha.title()}')
        # print(folha.title())
        pla = load_workbook(path)[folha]
        mostrar_conteudo_planilha(pla)
