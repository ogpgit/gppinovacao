from openpyxl import load_workbook, Workbook, open
from datetime import date


def manipularArqXl(arquivo):
    """
    Editar e alterar uma planilha excel

    :param arquivo: str com o nome do arquivo excel
    :return: str de "sucesso" ou "cancelado"

    """
    arquivo = f'G:\AmbientePython\PycharmProjects\gestaocontratos\{arquivo}'

    ws = Workbook()
    print(type(ws))
    ws1 = open(arquivo)
    print(type(ws1))

    #workbook = load_workbook('----------/dataset.xlsx')
    #sheet = workbook.active
    #row_count = sheet.max_row
    #for i in range(row_count):
    #    print(sheet.cell(row=i, column=2).value)
    sheet = ws1.active
    row_count = sheet.max_row
    print(row_count)
    for i in range(row_count):
        print(i)
        print(sheet.cell(row=i+1, column=1).value)


    return 'Sucesso'


if __name__ == '__main__':
    if manipularArqXl('sample.xlsx') == 'Sucesso':
        print('hello')
    else:
        print('Arquivo não existe')
